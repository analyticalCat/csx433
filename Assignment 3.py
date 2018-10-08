import decimal as dcm
import math
import struct
from array import array

import matplotlib.pyplot as plt
import numpy as np
import numpy.linalg as LA
import pandas as pd
from pylab import colorbar, pcolor, show, xticks, yticks


def readExcelSheet1(excelfile):
    from pandas import read_excel
    return (read_excel(excelfile)).values

def readExcelRange(excelfile,sheetname="Sheet1",startrow=1,endrow=1,startcol=1,endcol=1):
    from pandas import read_excel
    values=(read_excel(excelfile, sheetname,header=None)).values
    return values[startrow-1:endrow,startcol-1:endcol]

def readExcel(excelfile,**args):
    if args:
        data=readExcelRange(excelfile,**args)
    else:
        data=readExcelSheet1(excelfile)
    if data.shape==(1,1):
        return data[0,0]
    elif (data.shape)[0]==1:
        return data[0]
    else:
        return data

def writeExcelData(x,excelfile,sheetname,startrow,startcol):
    from pandas import DataFrame, ExcelWriter
    from openpyxl import load_workbook
    df=DataFrame(x)
    book = load_workbook(excelfile)
    writer = ExcelWriter(excelfile,  engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, sheet_name=sheetname,startrow=startrow-1, startcol=startcol-1, header=False, index=False)
    writer.save()
    writer.close()

def writeExcelDatabyRow(x,excelfile,sheetname,startrow,startcol):
    from pandas import DataFrame, ExcelWriter
    from openpyxl import load_workbook
    df=DataFrame(x)
    df=df.transpose()
    book = load_workbook(excelfile)
    writer = ExcelWriter(excelfile,  engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, sheet_name=sheetname,startrow=startrow-1, startcol=startcol-1, header=False, index=False)
    writer.save()
    writer.close()
    


def getSheetNames(excelfile):
    from pandas import ExcelFile
    return (ExcelFile(excelfile)).sheet_names

excelfile="c:\\Users\\jess_\\Downloads\\Assignment_3_ Submission_Jess.xlsx"

sheets=getSheetNames(excelfile);sheets
d=784
np.random.seed(0)
# Results, Scatter Plot, Observations

def load_mnist(dataset="training", selecteddigits=range(10), path=r'c:\\Users\\jess_\\Downloads\\\MNIST'):

    #Check training/testing specification. Must be "training" (default) or "testing"
    if dataset == "training":
        fname_digits = path + '\\' + 'train-images.idx3-ubyte'
        fname_labels = path + '\\' + 'train-labels.idx1-ubyte'
    elif dataset == "testing":
        fname_digits = path + '\\' + 't10k-images.idx3-ubyte'
        fname_labels = path + '\\' + 't10k-labels.idx1-ubyte'
    else:
        raise ValueError("dataset must be 'testing' or 'training'")
        
        
    #Import digits data
    digitsfileobject = open(fname_digits, 'rb')
    magic_nr, size, rows, cols = struct.unpack(">IIII", digitsfileobject.read(16))
    digitsdata = array("B", digitsfileobject.read())
    digitsfileobject.close()

    #Import label data
    labelsfileobject = open(fname_labels, 'rb')
    magic_nr, size = struct.unpack(">II", labelsfileobject.read(8))
    labelsdata=array("B",labelsfileobject.read())
    labelsfileobject.close()
    
    #Find indices of selected digits
    indices=[k for k in range(size) if labelsdata[k] in selecteddigits]
    N=len(indices)
    
    #Create empty arrays for X and T
    X = np.zeros((N, rows*cols), dtype=np.uint8)
    T = np.zeros((N, 1), dtype=np.uint8)
    
    #Fill X from digitsdata
    #Fill T from labelsdata
    for i in range(N):
        X[i] = digitsdata[indices[i]*rows*cols:(indices[i]+1)*rows*cols]
        T[i] = labelsdata[indices[i]]
    
    return X,T


def vectortoimg(v,show=True):
    plt.imshow(v.reshape(28, 28),interpolation='None', cmap='gray')
    plt.axis('off')
    if show:
        plt.show()

negdigit=5
posdigit=6
X, T = load_mnist(dataset="training",selecteddigits=[negdigit,posdigit])

# OTHER WAYS OF CALLING load_mnist
# X, T = load_mnist(dataset="training",selecteddigits=[5,6])
# X, T = load_mnist() #Loads ALL digits of training data
# X, T = load_mnist(dataset="testing",selecteddigits=[1,2,7])

print("Checking shape of matrix:", X.shape)
print("Checking min/max values:",(np.amin(X),np.amax(X)))
print("Checking unique labels in T:",list(np.unique(T)))

# Checking shape of matrix: (11339, 784)
# Checking min/max values: (0, 255)
# Checking unique labels in T: [5, 6]

print("Checking one training vector by plotting image:")
#vectortoimg(X[-20])

#count of 5 and 6
print(pd.value_counts(T[:,0]))

# print("Checking multiple training vectors by plotting images.\nBe patient:")
# plt.close('all')
# fig = plt.figure()
# nrows=10
# ncols=10
# for row in range(nrows):
#     for col in range(ncols):
#         plt.subplot(nrows, ncols, row*ncols+col + 1)
#         vectortoimg(X[np.random.randint(len(T))],show=False)
# plt.show()

#XZCVP
#the mean vector and two eigen vectors must be calcualted and entered into excel.

mu=np.mean(X, axis=0)
#vectortoimg(mu)
writeExcelDatabyRow(mu, excelfile, 'Results',2,2)
#print ('mean vectors') 
#writeExcelDatabyRow(mu, excelfile2, 'Sheet1',2,2)

Z=np.zeros((len(X), len(X[0])), dtype=np.uint8)
Z=X-mu
print("Checking min/max values:",(np.amin(Z),np.amax(Z)))

C=np.zeros((len(X[0]), len(X[0])), dtype=np.uint8)
#C=np.cov(Z,rowvar=False)  
C=np.cov(Z.T) 
def vectortoimg_new(v,show=True):
    plt.imshow(v.reshape(784, 784),interpolation='None', cmap='gray')
    plt.axis('off')
    if show:
        plt.show()

vectortoimg_new(C)

#validate C
for (x,y),value in np.ndenumerate(C):
    if (value!=C[y,x]): 
        print ('covariance wrong',x,y,value)
    if x==y and value<0: print ('covariance wrong', x,y,value)



EE, V=LA.eigh(C); 
#print (EE,'\n\n', EV)
#verify V
row=V[0,:];col=V[:,0]



#print 'validate V'
#print np.dot(C,row)-(EE[0]*row)
#print np.dot(C,col)-(EE[0]*col).  Col is the eigen vector
#Set the row as the eigenvector
EE=np.flipud(EE); V=np.flipud(V.T)
#row=EV[0,:]
#print np.dot(C,row)-(EE[0]*row)
#print LA.norm(V[0]);print LA.norm(V[1])
#print np.dot(V[0,:], V[1,:])
vectortoimg(V[0])
vectortoimg(V[1])

#writeExcelDatabyRow(V[0], excelfile,'Results',3,2)
#writeExcelDatabyRow(V[1], excelfile,'Results',4,2)

#Principal component
P=np.dot(Z,V.T)
#print P
print(np.mean(P[:,0:2], axis=0))

R=np.dot(P,V)
#print (R-Z)  R-Z is almost 0 means R is recovered

Xrec=R+mu
print(Xrec-X)  #The result is [0,,,0] means we have recovered X

#Dimension reduction
Xrec1=(np.dot(P[:,0:1],V[0:1,:]))+mu
print( 'xrec1', Xrec1)
print (np.amin(Xrec1), np.amax(Xrec1))

Xrec2=(np.dot(P[:,0:2],V[0:2,:]))+mu;print ('xrec2',Xrec2)
print( np.amin(Xrec2), np.amax(Xrec2))

Xrec3=(np.dot(P[:,0:3],V[0:3,:]))+mu;print ('xrec3',Xrec3)
print( np.amin(Xrec3), np.amax(Xrec3))

# vectortoimg(Xrec1[20])
# vectortoimg(Xrec2[20])
vectortoimg(Xrec3[20])
# so far so good.  check the rest

# For best effect, points should not be drawn in sequence but in random order
np.random.seed(0)
randomorder=np.random.permutation(np.arange(len(T)))
randomorder=np.arange(len(T))

# Set colors
cols=np.zeros((len(T),4))     # Initialize matrix to hold colors
cols[T.ravel()==negdigit,:]=[1,0,0,0.25] # Negative points are red (with opacity 0.25)
cols[T.ravel()==posdigit,:]=[0,1,0,0.25] # Positive points are green (with opacity 0.25)

# Draw scatter plot
fig = plt.figure()
ax = fig.add_subplot(111, facecolor='black')
ax.scatter(P[randomorder,1],P[randomorder,0],s=5,linewidths=0,facecolors=cols[randomorder,:],marker="o")
ax.set_aspect('equal')
plt.gca().invert_yaxis()
plt.show()

#The point clouds show the original graph, doesn't look like it has been moved and rotated.

#Histogram classifier building
B=int(math.log(len(X), 2))+1
B=25 #according to the homework
H1=np.zeros((B,B),dtype=np.uint8)
H2=np.zeros((B,B),dtype=np.uint8)

#find the min and max for training data p1 and p2
pmin1=np.amin(P[:,0])
pmax1=np.amax(P[:,0])
pmin2=np.amin(P[:,1])
pmax2=np.amax(P[:,1])
print( 'training set mins and maxes pmin1,pmax1,pmin2,pmax2',pmin1,pmax1,pmin2,pmax2)

#Build H1=5 and H2=6
N=len(P)
i=0
while i<N:
    r=int((B-1)*(P[i,0]-pmin1 )/(pmax1-pmin1))
    c=int((B-1)*(P[i,1]-pmin2)/(pmax2-pmin2))
    #print 'r and c'; print (r,c)
    if T[i,0]==negdigit:
        H1[r,c]=H1[r,c]+1
    else:
        H2[r,c]=H2[r,c]+1
    i+=1


#apply the 2d histogram classifier
def build2DHistogramclassifier(x,mu,V,B,pmin1,pmin2,pmax1,pmax2,H1,H2):
    z=x-mu
    p=np.dot(z,V[0:2].T)
    r=int((B-1)*(p[0]-pmin1 )/(pmax1-pmin1))
    c=int((B-1)*(p[1]-pmin2)/(pmax2-pmin2))
    if H1[r,c]+H2[r,c] ==0: 
        print( 'undecided')
        resultNeg=0; resultPos=0
    else:
        resultNeg=float(H1[r,c])/(float(H1[r,c])+float(H2[r,c]))
        resultPos=float(H2[r,c])/(float(H1[r,c])+float(H2[r,c]))
        #print ('probability1 and 2', result1, result2)
        #1 is negative, 2 is positive
    return resultNeg, resultPos  


x=X[-20]
result1,result2=build2DHistogramclassifier(x,mu,V,B,pmin1,pmin2,pmax1,pmax2,H1,H2)
print('the probability of X[-20] being positive is ',result2)
x=X[20]
result1,result2=build2DHistogramclassifier(x,mu,V,B,pmin1,pmin2,pmax1,pmax2,H1,H2)
print('the probability of X[20] being positive is ',result2)
 
def buildPDF(x, mu, sigma):
    d=np.alen(mu)
    dfact1=(2*np.pi)**d
    dfact2=np.linalg.det(sigma)
    dfact3=1/np.sqrt(dfact1*dfact2)

    z=x-mu
    isigma=np.linalg.inv(sigma)
    afactor1=np.dot(z, isigma)
    afactor2=(-0.5)*np.dot(afactor1, z.T)
    #QQ=dfact3*np.exp(-0.5*np.einsum('ij,jk,ik->i',z,isigma,z))
    Q=dfact3*np.exp(afactor2)
    
    return Q

def pdf(x,mu,sigma):

    d=np.alen(mu)

    dfact1=(2*np.pi)**d

    dfact2=np.linalg.det(sigma)

    fact=1/np.sqrt(dfact1*dfact2)

    xc=x-mu

    isigma=np.linalg.inv(sigma)

    return fact*np.exp(-0.5*np.einsum('ij,jk,ki->i',xc,isigma,xc))

negmean=np.mean(P[T[:,0]==negdigit,0:2],axis=0)
posmean=np.mean(P[T[:,0]==posdigit,0:2],axis=0)
print ('neg mean vector and pos mean vector',negmean,posmean)

Nneg=0;Npos=0
Sumneg=0.0;Sumneg2=0.0;Sumpos=0.0;Sumpos2=0.0
for i, row in enumerate(P):
    if T[i,0]==negdigit:
        Nneg+=1
        Sumneg+=P[i,0]
    else:
        Npos+=1
        Sumpos+=P[i,0]
print(Sumneg/Nneg, Sumpos/Npos)
    
            

covneg=np.cov(P[T[:,0]==negdigit,0:2],rowvar=False)
covpos=np.cov(P[T[:,0]==posdigit,0:2],rowvar=False)
print ('covariance',covneg,covpos)

#writeExcelData(H1,excelfile,'Results',46,2)
#writeExcelData(H2,excelfile,'Results',20,2)


#use two examples to show result
xp=X[-20] #6
#buildBayesianClassifier(xp, mu, V, 5918,5421,posmean,negmean,covneg,covpos)
QP=buildPDF(P[-20,0:2], posmean, covpos)
QN=buildPDF(P[-20,0:2], negmean, covneg)
print(QP/(QP+QN))
xn=X[20]
QP=buildPDF(P[20,0:2], posmean, covpos)
QN=buildPDF(P[20,0:2], negmean, covneg)
print (QP/(QP+QN))

# zp=xp-mu
# pp=np.dot(zp,V[0:2].T)
# rzp=np.dot(pp,V[0:2])
# xrecp=rzp+mu
# writeExcelDatabyRow(xp, excelfile, 'Results',74,2)
# writeExcelDatabyRow(zp, excelfile, 'Results',75,2)
# writeExcelDatabyRow(pp, excelfile, 'Results',76,2)
# writeExcelDatabyRow(rzp, excelfile, 'Results',77,2)
# writeExcelDatabyRow(xrecp, excelfile, 'Results',78,2)
# vectortoimg(xrecp)
#use the formula to calculate probability of class labels.




#xn=X[20] #5 negative
# zn=xn-mu
# pn=np.dot(zn, V[0:2].T)
# rzn=np.dot(pn,V[0:2])
# xrecn=rzn+mu
# vectortoimg(xrecn)
# writeExcelDatabyRow(xn, excelfile, 'Results',80,2)
# writeExcelDatabyRow(zn, excelfile, 'Results',81,2)
# writeExcelDatabyRow(pn, excelfile, 'Results',82,2)
# writeExcelDatabyRow(rzn, excelfile, 'Results',83,2)
# writeExcelDatabyRow(xrecn, excelfile, 'Results',84,2)

#calculate the training accuracy.
#first, training accuracy using hitogram
wrong=0
for i,x in enumerate(X):
    result1,result2=build2DHistogramclassifier(x,mu,V,B,pmin1,pmin2,pmax1,pmax2,H1,H2)
    if result1==result2: wrong+=1
    if result1<result2 and T[i]==5: wrong+=1
    if result1>result2 and T[i]==6: wrong+=1
    
trainingaccuracy=(len(T)-wrong)/len(T)
print ('histogram training accuracy:', trainingaccuracy)

#second, training accuracy using bayesian
wrong=0
for i,p in enumerate(P):
    result1=buildPDF(p[0:2], negmean, covneg)
    result2=buildPDF(p[0:2], posmean, covpos)
    if result1==result2: wrong+=1
    if result1<result2 and T[i]==5: wrong+=1
    if result1>result2 and T[i]==6: wrong+=1

trainingaccuracy=(len(T)-wrong)/len(T)
print ('Bayesian training accuracy:', trainingaccuracy)

#from the professor
#def build2dhistogram(P,T,B,mn,mx) mn and mx are min vector and max vector
#def Histogram2dclassifier(queries,mn,mx,Hn,Hp,labeln,labelp):
#def bayesian2dclarrifier(queries,Nn,Np,nun,mup,cn,cp,lablen,lablep)

    # for i,q in enumberate(queries):
    #         countn[i]=factorn*exp....

# T=labels.flatten()
# unituq(T,return_countes=true)
# mu=mean(X,axis=0)
# Z=X=muC=cov(Z,rowvar=false)
# E=flipud(E)
# V=flipud(V.T)
# V=V[:2]
# P=dot(Z,V.T)
# R=dot(P,V)
# Xrec=R+mu

# Vectortoimg(C, width=84,size=5)  #try this again. 
# VC-EV=0
# [max9abs(dot(C,V[0,:])-E[0]*V)]v[0,:])),max(abs(dot(C,V[1,...])))
# [out] close to be 0
# dot(v[0],v[1]) almost=0
# vectortoimg (V[0,:],V[1,:]) should show image of two digits.
# you can find the accuracy programmtically. T, blabels
# The histogram is more accurate in this case because the distribution is not necessary normal. histogram does not assume that.


